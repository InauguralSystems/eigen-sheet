#!/usr/bin/env python3
"""Differential oracle: eigen-sheet vs a real spreadsheet (LibreOffice Calc).

For the numeric operations every spreadsheet agrees on — arithmetic with
precedence, cell references, SUM over a range — "what right looks like" is not
ours to assert. It's what a real, used spreadsheet computes. So we write the
SAME cells and formulas into an .xlsx (fullCalcOnLoad, so no stale cached
values), convert it to CSV with headless LibreOffice Calc, and float-diff its
computed values against eigen-sheet's recalc. LibreOffice's numbers are the
reference.

Error semantics (#CYCLE, div-by-zero) and cursor/selection behavior differ
across engines, so they are NOT checked here — they live in the model oracle.
This oracle owns the numeric consensus only.

Skips (exit 2) when openpyxl or a working LibreOffice *Calc* is unavailable —
e.g. a box with libreoffice-core but not libreoffice-calc. CI installs both
and treats a skip as failure, so the check is never silently dropped there.
"""
import csv, glob, os, re, shutil, subprocess, sys, tempfile

EIGS = os.environ.get("EIGENSCRIPT", "eigenscript")
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOFFICE = next((b for b in ("libreoffice", "soffice", "localc") if shutil.which(b)), None)

# Test grid: address -> raw cell content (numbers and formulas). Numeric only.
CELLS = {
    "A1": "5",    "A2": "3",    "A3": "8",     "A4": "-2",
    "B1": "=A1+A2",         # 8
    "B2": "=A1*A2-A4",      # 17
    "B3": "=(A1+A2)*A3",    # 64
    "B4": "=A1/A2",         # 1.666...
    "C1": "=SUM(A1:A4)",    # 14
    "C2": "=SUM(A1:A3)/2",  # 8
    "C3": "=B1+C1",         # 22   (formula referencing formulas)
    "C4": "=-A1+A3*2",      # 11   (unary minus + precedence)
    "D1": "=SUM(A1:C1)",    # 27   (A1+B1+C1 = 5+8+14)
    "E1": "=AVERAGE(A1:A4)",             # 3.5  (5+3+8-2)/4
    "E2": "=MIN(A1:A4)",                 # -2
    "E3": "=MAX(A1:A4)",                 # 8
    "E4": "=IF(A1>A2,100,200)",          # 100  (5 > 3)
    "F1": "=IF(A1<A2,10,20)",            # 20   (5 < 3 false)
    "F2": "=IF(SUM(A1:A4)>10,MAX(A1:A4),MIN(A1:A4))",  # 8 (14>10 -> MAX)
    "F3": "=IF(A2=3,A3*2,0)",            # 16   (equality)
    # ranges with GAPS and TEXT — aggregates must skip them, like Calc.
    "G1": "10", "G3": "20",              # G2 deliberately empty
    "H1": "=AVERAGE(G1:G3)",             # 15  ((10+20)/2, empty G2 skipped)
    "H2": "=MIN(G1:G3)",                 # 10  (not 0)
    "H3": "=MAX(G1:G3)",                 # 20
    "H4": "=SUM(G1:G3)",                 # 30
    "I1": "note", "I2": "7",             # I1 is text
    "J1": "=AVERAGE(I1:I2)",             # 7   (text I1 skipped)
    "J2": "=MIN(I1:I2)",                 # 7
    # ---- math & rounding library (#1) ----
    "K1": "=ABS(A4)",                    # 2    (A4 = -2)
    "K2": "=SIGN(A4)",                   # -1
    "K3": "=SIGN(A1)",                   # 1
    "K4": "=INT(B4)",                    # 1    (B4 = 1.666..)
    "L1": "=ROUND(B4,2)",                # 1.67
    "L2": "=ROUNDUP(B4,1)",              # 1.7
    "L3": "=ROUNDDOWN(B4,2)",            # 1.66
    "L4": "=TRUNC(B4)",                  # 1
    "O1": "=ROUND(1234.5678,-2)",        # 1200 (negative digits)
    "M1": "=MOD(A3,A2)",                 # 2    (8 mod 3)
    "M2": "=MOD(A4,A2)",                 # 1    (-2 mod 3 -> sign of divisor)
    "M3": "=POWER(A2,A1)",               # 243  (3^5)
    "M4": "=A2^A1",                      # 243  (^ operator)
    "N1": "=SQRT(A3)",                   # 2.828..
    "N2": "=CEILING(A3,A2)",             # 9    (8 up to a multiple of 3)
    "N3": "=FLOOR(A3,A2)",               # 6
    "N4": "=PRODUCT(A1:A3)",             # 120  (5*3*8)
    "O2": "=-A1^2",                      # 25   (unary minus binds tighter: (-5)^2)
    "O3": "=2^3^2",                      # 64   (^ is left-associative in Calc)
    "O4": "=TRUNC(-A1/A2)",              # -1   (toward zero)
    # ---- statistical library (#2) ----
    "P1": "4", "P2": "4", "P3": "7", "P4": "4",
    "Q1": "=COUNT(A1:A4)",               # 4
    "Q2": "=COUNT(G1:G3)",               # 2    (empty G2 skipped)
    "Q3": "=COUNTA(G1:G3)",              # 2
    "Q4": "=COUNTA(I1:I2)",              # 2    (text counts)
    "R1": "=COUNTBLANK(G1:G3)",          # 1    (G2)
    "R2": "=MEDIAN(A1:A4)",              # 4    ([-2,3,5,8] -> (3+5)/2)
    "R3": "=MEDIAN(A1:A3)",              # 5    ([3,5,8])
    "R4": "=MODE(P1:P4)",                # 4    (most frequent)
    "T1": "=MEDIAN(P1:P4)",              # 4    ([4,4,4,7] -> (4+4)/2)
    "S1": "=VAR(A1:A4)",                 # 17.6667  (sample, n-1)
    "S2": "=VARP(A1:A4)",                # 13.25    (population, n)
    "S3": "=STDEV(A1:A4)",               # 4.20317  (sqrt sample var)
    "S4": "=STDEVP(A1:A4)",              # 3.64005  (sqrt pop var)
    # ---- absolute / mixed references (#9): evaluate identically to relative ----
    "U1": "=$A$1+$A2+A$3",               # 16   (5+3+8, anchors are eval no-ops)
    "U2": "=SUM($A$1:$A$3)",             # 16
    # ---- logical functions + IF short-circuit (#3) ----
    # booleans wrapped so results stay numeric (Calc exports bare TRUE/FALSE as text)
    "V1": "=IF(A1>A2,100,A1/0)",         # 100  (untaken 1/0 NOT evaluated)
    "V2": "=IF(A1<A2,A1/0,200)",         # 200  (untaken 1/0 NOT evaluated)
    "V3": "=IF(AND(A1>0,A2>0),1,0)",     # 1
    "V4": "=IF(AND(A1>0,A4>0),1,0)",     # 0    (A4 = -2)
    "W1": "=IF(OR(A4>0,A1>0),1,0)",      # 1
    "W2": "=IF(OR(A4>0,A4>10),1,0)",     # 0
    "W3": "=IF(NOT(A1>A2),1,0)",         # 0    (A1>A2 true -> NOT false)
    "W4": "=IF(XOR(A1>0,A4>0),1,0)",     # 1    (exactly one true)
    "X6": "=IF(XOR(A1>0,A2>0),1,0)",     # 0    (both true -> even)
    "X7": "=AND(A1>0,A2>0)*10",          # 10   (bare AND coerces 1)
    "X8": "=TRUE()+FALSE()",             # 1
    "X9": "=IFS(A1>10,1,A1>3,2,A1>0,3)", # 2    (first true branch: 5>3)
    "Y4": "=IFS(A4>0,1,A4<0,2,A4=0,3)",  # 2    (A4<0)
    "Y5": "=SWITCH(A2,1,10,3,30,99)",    # 30   (A2=3 matches)
    "Y6": "=SWITCH(A1,1,10,2,20,99)",    # 99   (no match -> default)
    # ---- lookup & reference (#4) ----
    # a sorted numeric table: keys Z1:Z4, values AA1:AA4
    "Z1": "10", "Z2": "20", "Z3": "30", "Z4": "40",
    "AA1": "100", "AA2": "200", "AA3": "300", "AA4": "400",
    # a horizontal table: keys AB1:AD1, values AB2:AD2
    "AB1": "1", "AC1": "2", "AD1": "3",
    "AB2": "11", "AC2": "22", "AD2": "33",
    "AE1": "=VLOOKUP(30,Z1:AA4,2,0)",    # 300  (exact)
    "AE2": "=VLOOKUP(25,Z1:AA4,2,1)",    # 200  (approx: largest key <=25 is 20)
    "AE3": "=VLOOKUP(25,Z1:AA4,2)",      # 200  (approx is the default)
    "AE4": "=HLOOKUP(2,AB1:AD2,2,0)",    # 22   (exact, horizontal)
    "AF1": "=MATCH(30,Z1:Z4,0)",         # 3    (exact position)
    "AF2": "=MATCH(25,Z1:Z4,1)",         # 2    (approx position)
    "AF3": "=INDEX(AA1:AA4,3)",          # 300  (1-D column vector)
    "AF4": "=INDEX(Z1:AA4,2,2)",         # 200  (2-D row,col)
    "AG1": "=ROWS(Z1:AA4)",              # 4
    "AG2": "=COLUMNS(Z1:AA4)",           # 2
    "AG3": "=ROW(Z3)",                   # 3
    "AG4": "=COLUMN(AA1)",               # 27   (AA)
    "AH3": "=LOOKUP(25,Z1:Z4,AA1:AA4)",  # 200  (vector form, approx)
    "AH4": "=MATCH(40,Z1:Z4,0)",         # 4
}


def col_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def addr_parts(a):
    m = re.match(r"([A-Z]+)([0-9]+)", a)
    return m.group(1), int(m.group(2))


def eigs_values():
    """eigen-sheet's computed value for every formula cell: {addr: float}."""
    body = "\n".join('sheet.set_cell of [s, "%s", "%s"]' % (a, raw)
                     for a, raw in CELLS.items())
    prints = "\n".join(
        'print of ("%s\\t" + (str of (sheet.get of [s, "%s"])))' % (a, a)
        for a, raw in CELLS.items() if raw.startswith("="))
    prog = ("import sheet\ns is sheet.new_sheet of null\n%s\nsheet.recalc of s\n%s\n"
            % (body, prints))
    tmp = tempfile.mkdtemp()
    try:
        md = os.path.join(tmp, "eigs_modules", "sheet"); os.makedirs(md)
        shutil.copy(os.path.join(REPO, "sheet.eigs"), os.path.join(md, "sheet.eigs"))
        shutil.copy(os.path.join(REPO, "eigs.json"), os.path.join(md, "eigs.json"))
        app = os.path.join(tmp, "app.eigs"); open(app, "w").write(prog)
        r = subprocess.run([EIGS, app], cwd=tmp, capture_output=True, text=True, timeout=60)
        if r.returncode != 0:
            raise RuntimeError(r.stdout + r.stderr)
        out = {}
        for line in r.stdout.splitlines():
            if "\t" in line:
                a, v = line.split("\t")
                out[a] = float(v)
        return out
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# "Future functions" added after the original OOXML spec are stored in .xlsx
# with an _xlfn. prefix; openpyxl writes the bare name, which LibreOffice (and
# Excel) then reject as #NAME?. eigen-sheet always gets the PLAIN formula — only
# the .xlsx handed to LibreOffice carries the prefix, so both compute the same fn.
_XLFN = ("XOR", "IFS", "SWITCH")


def _ooxml(formula):
    import re
    for name in _XLFN:
        formula = re.sub(r"\b" + name + r"\(", "_xlfn." + name + "(", formula)
    return formula


def write_xlsx(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for a, raw in CELLS.items():
        if raw.startswith("="):
            ws[a] = _ooxml(raw)
        else:
            try:
                ws[a] = float(raw)
            except ValueError:
                ws[a] = raw   # text literal
    wb.calculation.fullCalcOnLoad = True   # force Calc to recompute on load
    wb.save(path)


def calc_values():
    """LibreOffice's computed value for every formula cell: {addr: float}, or None to skip."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("SKIP: openpyxl not installed")
        return None
    tmp = tempfile.mkdtemp()
    try:
        xlsx = os.path.join(tmp, "grid.xlsx"); write_xlsx(xlsx)
        profile = "file://" + os.path.join(tmp, "profile")
        subprocess.run([SOFFICE, "--headless", "--norestore", "--nolockcheck",
                        "-env:UserInstallation=" + profile,
                        "--convert-to", "csv", "--outdir", tmp, xlsx],
                       capture_output=True, timeout=180)
        csvs = glob.glob(os.path.join(tmp, "*.csv"))
        if not csvs:
            print("SKIP: LibreOffice produced no CSV (is libreoffice-calc installed?)")
            return None
        grid = list(csv.reader(open(csvs[0], newline="")))
        out = {}
        for a, raw in CELLS.items():
            if not raw.startswith("="):
                continue
            col, row = addr_parts(a)
            out[a] = float(grid[row - 1][col_to_num(col) - 1])
        return out
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    if not SOFFICE:
        print("SKIP: no libreoffice/soffice found — external oracle needs one")
        sys.exit(2)
    print("reference engine: %s" % SOFFICE)
    ref = calc_values()
    if ref is None:
        sys.exit(2)
    mine = eigs_values()
    failures = 0
    for a in sorted(CELLS):
        if not CELLS[a].startswith("="):
            continue
        m, r = mine.get(a), ref.get(a)
        if m is None or r is None or abs(m - r) > 1e-9:
            failures += 1
            print("FAIL %-4s eigen-sheet=%r  libreoffice=%r  (%s)" % (a, m, r, CELLS[a]))
        else:
            print("PASS %-4s %-14s = %g" % (a, CELLS[a], r))
    if failures:
        print("\n%d divergence(s) from LibreOffice" % failures)
        sys.exit(1)
    print("\neigen-sheet matches LibreOffice on every numeric formula")


if __name__ == "__main__":
    main()
