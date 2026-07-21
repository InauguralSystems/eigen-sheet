#!/usr/bin/env python3
"""String differential oracle: eigen-sheet vs a real spreadsheet (LibreOffice).

The companion to tests/diff_vs_calc.py, for the TEXT side of the language:
string literals, the & concatenation operator, the text functions
(LEN/UPPER/LOWER/TRIM/LEFT/RIGHT/MID/CONCATENATE), text comparison in IF, and
number<->text coercion. "What right looks like" is again not ours to assert —
it is what LibreOffice Calc computes. We write the SAME cells into an .xlsx,
convert it to CSV with headless Calc, and byte-compare its cell TEXT against
eigen-sheet's display() output.

Only well-defined results live here. Error tokens (#NAME?, #VALUE!, #DIV/0!)
differ across engines, so — exactly as in the numeric oracle — they are checked
in the model oracle (tests/test_smoke.sh), not against the external reference.

Skips (exit 2) when openpyxl or a working LibreOffice *Calc* is unavailable.
CI installs both and treats a skip as failure, so the check is never silently
dropped there.
"""
import csv, glob, os, re, shutil, subprocess, sys, tempfile

EIGS = os.environ.get("EIGENSCRIPT", "eigenscript")
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOFFICE = next((b for b in ("libreoffice", "soffice", "localc") if shutil.which(b)), None)

# Test grid: address -> raw cell content. Text literals, string formulas, and
# formulas that MIX text and numbers. Every value here is one LibreOffice
# computes deterministically to a plain string (no error tokens).
CELLS = {
    # text-literal inputs the formulas below reference
    "B1": "hi",  "B2": "yo",  "B3": "5",  "B4": "World",
    # text functions
    "A1": '=LEFT("hello",2)',           # he
    "A2": '=RIGHT("hello",2)',          # lo
    "A3": '=MID("hello",2,3)',          # ell
    "A4": '=LEN("hello")',              # 5   (a number result)
    "A5": '=UPPER("abc")',              # ABC
    "A6": '=LOWER("AbC")',              # abc
    "A7": '=TRIM("  a   b  ")',         # a b (internal run collapsed)
    "A8": '=LEFT("hi",5)',              # hi  (n clamped to length)
    "A9": '=MID("hello",2,100)',        # ello
    # concatenation + number<->text coercion
    "A10": '=5&"x"',                    # 5x
    "A11": '="a"&5&"b"',                # a5b
    "A12": '=B1&B2',                    # hiyo
    "A13": '=B1&" "&B4',                # hi World
    "A14": '=B3&B3',                    # 55  (text join, not 10)
    "A15": '="Row "&(1+1)',             # Row 2  (arith inside concat)
    "A16": '=CONCATENATE("a","b","c")', # abc
    "A17": '=CONCATENATE(B1,"-",B4)',   # hi-World
    # text comparison in IF (case-insensitive equality, string ordering)
    "A18": '=IF("a"="A","EQ","NE")',    # EQ  (case-insensitive)
    "A19": '=IF("apple"<"banana","1st","2nd")',  # 1st
    "A20": '=IF(B1="HI","match","no")', # match (case-insensitive vs cell)
    "A21": '=UPPER(LEFT("hello",3))',   # HEL (nested)
    # ---- text library extend (#5) ----
    "C1": '=FIND("l","hello")',            # 3   (1-based, case-sensitive)
    "C2": '=SEARCH("L","hello")',          # 3   (case-insensitive)
    "C3": '=SUBSTITUTE("a-b-c","-","+")',  # a+b+c  (all)
    "C4": '=SUBSTITUTE("a-b-c","-","+",2)',# a-b+c  (2nd only)
    "C5": '=REPLACE("abcdef",2,3,"XY")',   # aXYef
    "C6": '=REPT("ab",3)',                 # ababab
    "C7": '=PROPER("hello world")',        # Hello World
    "C8": '=IF(EXACT("aB","aB"),"y","n")', # y
    "C9": '=IF(EXACT("ab","aB"),"y","n")', # n   (case-sensitive)
    "C10": '=CHAR(65)',                    # A
    "C11": '=CODE("A")',                   # 65
    "C12": '=VALUE("42")',                 # 42
    "C13": '=VALUE("3.5")+1',              # 4.5 (VALUE -> number)
    "C14": '=TEXTJOIN("-",1,"a","","b")',  # a-b  (ignore empty)
    "C15": '=TEXTJOIN(",",0,"a","","b")',  # a,,b (keep empty)
    "C16": '=PROPER("mIxEd cASE 3rd")',    # Mixed Case 3Rd
    # ---- first-class booleans (#13): display as TRUE/FALSE ----
    "D1": '=1>0',                          # TRUE  (comparison operator)
    "D2": '=2<1',                          # FALSE
    "D3": '=TRUE()',                       # TRUE
    "D4": '=FALSE()',                      # FALSE
    "D5": '=AND(1>0,2>0)',                 # TRUE
    "D6": '=OR(1>2,3>4)',                  # FALSE
    "D7": '=NOT(1>0)',                     # FALSE
    "D8": '=EXACT("aB","aB")',             # TRUE
    "D9": '=XOR(1>0,2>0)',                 # FALSE (both true -> even)
    "D10": '=TRUE()&"!"',                  # 1!  (concat coerces a bool to "1"/"0", NOT its display text)
    "D11": '=(5>3)&"/"&(1>9)',             # 1/0
    # ---- TEXT() number/date formatting (#23) ----
    "E1": '=TEXT(1234.5,"0.00")',          # 1234.50
    "E2": '=TEXT(1234.5,"#,##0")',         # 1,235  (rounds)
    "E3": '=TEXT(1234.5,"#,##0.00")',      # 1,234.50
    "E4": '=TEXT(0.1234,"0.00%")',         # 12.34%
    "E5": '=TEXT(0.1234,"0%")',            # 12%
    "E6": '=TEXT(1234.5,"$#,##0.00")',     # $1,234.50
    "E7": '=TEXT(-1234.5,"#,##0.00")',     # -1,234.50
    "E8": '=TEXT(5,"0.000")',              # 5.000
    "E9": '=TEXT(1234567,"#,##0")',        # 1,234,567
    "E10": '=TEXT(DATE(2020,1,5),"YYYY-MM-DD")',  # 2020-01-05
    "E11": '=TEXT(DATE(2020,1,5),"MM/DD/YYYY")',  # 01/05/2020
    "E12": '=TEXT(DATE(2020,3,9),"D/M/YYYY")',    # 9/3/2020
    "E13": '=TEXT(0,"0.00")',              # 0.00
    "E14": '=TEXT(2.5,"0")',               # 3 (half away)
}


def col_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def addr_parts(a):
    m = re.match(r"([A-Z]+)([0-9]+)", a)
    return m.group(1), int(m.group(2))


def eigs_str(s):
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def eigs_values():
    """eigen-sheet's display() string for every formula cell: {addr: str}."""
    body = "\n".join("sheet.set_cell of [s, %s, %s]" % (eigs_str(a), eigs_str(raw))
                     for a, raw in CELLS.items())
    prints = "\n".join(
        'print of (%s + "\\t" + (sheet.display of [s, %s]))' % (eigs_str(a), eigs_str(a))
        for a, raw in CELLS.items() if raw.startswith("="))
    prog = ("import sheet\ns is sheet.new_sheet of null\n%s\nsheet.recalc of s\n%s\n"
            % (body, prints))
    tmp = tempfile.mkdtemp()
    try:
        md = os.path.join(tmp, "eigs_modules", "sheet"); os.makedirs(md)
        shutil.copy(os.path.join(REPO, "sheet.eigs"), os.path.join(md, "sheet.eigs"))
        shutil.copy(os.path.join(REPO, "eigs.json"), os.path.join(md, "eigs.json"))
        app = os.path.join(tmp, "app.eigs"); open(app, "w").write(prog)
        r = subprocess.run([EIGS, app], cwd=tmp, capture_output=True, text=True, timeout=60)
        if r.returncode != 0:
            raise RuntimeError(r.stdout + r.stderr)
        out = {}
        for line in r.stdout.splitlines():
            if "\t" in line:
                a, v = line.split("\t", 1)
                out[a] = v
            elif line:   # empty display prints "ADDR\t" -> ["ADDR", ""]
                out[line] = ""
        return out
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# OOXML "future functions" are stored _xlfn.-prefixed; openpyxl writes the bare
# name and LibreOffice reads #NAME?. eigen-sheet gets the plain formula.
_XLFN = ("TEXTJOIN", "XOR")


def _ooxml(formula):
    for name in _XLFN:
        formula = re.sub(r"\b" + name + r"\(", "_xlfn." + name + "(", formula)
    return formula


def write_xlsx(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for a, raw in CELLS.items():
        if raw.startswith("="):
            ws[a] = _ooxml(raw)
        else:
            ws[a] = raw   # everything else is a text literal (incl "5")
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)


def calc_values():
    """LibreOffice's cell TEXT for every formula cell: {addr: str}, or None to skip."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("SKIP: openpyxl not installed")
        return None
    tmp = tempfile.mkdtemp()
    try:
        xlsx = os.path.join(tmp, "grid.xlsx"); write_xlsx(xlsx)
        profile = "file://" + os.path.join(tmp, "profile")
        subprocess.run([SOFFICE, "--headless", "--norestore", "--nolockcheck",
                        "-env:UserInstallation=" + profile,
                        "--convert-to", "csv", "--outdir", tmp, xlsx],
                       capture_output=True, timeout=180)
        csvs = glob.glob(os.path.join(tmp, "*.csv"))
        if not csvs:
            print("SKIP: LibreOffice produced no CSV (is libreoffice-calc installed?)")
            return None
        grid = list(csv.reader(open(csvs[0], newline="")))
        out = {}
        for a, raw in CELLS.items():
            if not raw.startswith("="):
                continue
            col, row = addr_parts(a)
            try:
                out[a] = grid[row - 1][col_to_num(col) - 1]
            except IndexError:
                out[a] = "<oob>"
        return out
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    if not SOFFICE:
        print("SKIP: no libreoffice/soffice found — external oracle needs one")
        sys.exit(2)
    print("reference engine: %s" % SOFFICE)
    ref = calc_values()
    if ref is None:
        sys.exit(2)
    mine = eigs_values()
    failures = 0
    for a in sorted(CELLS, key=lambda s: (len(s), s)):
        if not CELLS[a].startswith("="):
            continue
        m, r = mine.get(a), ref.get(a)
        if m != r:
            failures += 1
            print("FAIL %-4s eigen-sheet=%r  libreoffice=%r  (%s)" % (a, m, r, CELLS[a]))
        else:
            print("PASS %-4s %-26s = %r" % (a, CELLS[a], r))
    if failures:
        print("\n%d divergence(s) from LibreOffice" % failures)
        sys.exit(1)
    print("\neigen-sheet matches LibreOffice on every string formula")


if __name__ == "__main__":
    main()
